import os
import requests
from datetime import datetime, timedelta

from PySide6.QtCore import Qt, QThread, Signal, QDate
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox,
    QLineEdit, QDateEdit, QTableWidget, QTableWidgetItem, QHeaderView,
    QAbstractItemView, QMessageBox, QFileDialog,
)
from PySide6.QtGui import QFont

from auth_utils import build_auth_config

RESOURCE_MAP = {
    "METAR": "metar",
    "SINOPTIK": "synoptic",
}

CURRENT_DIR = os.path.dirname(os.path.abspath(__file__))
AUTH_STATE_FILE = os.path.join(CURRENT_DIR, "auth_state.json")
BASE_URL = "https://bmkgsatu.bmkg.go.id/api/v21/monitoring/gts/{resource}/daily/date/{tanggal}"

FIELD_METADATA = {"station_id", "station_name", "region", "propinsi", "kabupaten", "jam_pengamatan"}


def load_config():
    if not os.path.exists(AUTH_STATE_FILE):
        raise Exception(
            f"File '{AUTH_STATE_FILE}' tidak ditemukan.\n"
            "Jalankan 'Update Sesi Login' (save_session.py) dulu untuk login ke BMKG."
        )
    return build_auth_config(AUTH_STATE_FILE)


def build_session(config):
    s = requests.Session()
    s.headers.update({
        "accept": "application/json",
        "accept-language": "en-US,en;q=0.9,id;q=0.8",
        "authorization": f"Bearer {config['bearer_token']}",
        "cookie": config["cookie"],
        "referer": "https://bmkgsatu.bmkg.go.id/monitoring/gts",
        "user-agent": config.get(
            "user_agent",
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/150.0.0.0 Safari/537.36 Edg/150.0.0.0",
        ),
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
    })
    return s


def fetch_daily(session, tanggal, balai, provinsi, resource="metar"):
    """tanggal: 'YYYY-MM-DD' | balai/provinsi: ID (int) | resource: 'metar'/'synoptic'"""
    url = BASE_URL.format(resource=resource, tanggal=tanggal)
    params = {"balai": balai, "provinsi": provinsi}
    resp = session.get(url, params=params, timeout=30)

    if resp.status_code == 401:
        raise Exception("401 Unauthorized -> bearer_token expired, ambil token baru dari browser.")
    if resp.status_code == 403:
        raise Exception("403 Forbidden -> cf_clearance (Cloudflare) expired, ambil cookie baru.")
    resp.raise_for_status()
    return resp.json()


def parse_station(field_list):
    """
    Mengubah 1 daftar field (list of dict {key, value, status, raw, ...})
    menjadi dict metadata + dict jam.

    Return: (meta: dict, jam_data: dict)
      meta   = {"station_id": ..., "station_name": ..., "region": ..., "propinsi": ..., "kabupaten": ...}
      jam_data = {"00:00": {"value": "1", "status": "ok"}, "00:30": {...}, ...}
    """
    meta = {}
    jam_data = {}

    for field in field_list:
        key = field.get("key")
        if key in FIELD_METADATA:
            meta[key] = field.get("value")
        else:
            jam_data[key] = {
                "value": field.get("value"),
                "status": field.get("status"),
                "raw": field.get("raw", []),
            }

    return meta, jam_data


def parse_response(json_data):
    """
    Mem-parse seluruh response API menjadi list of (meta, jam_data),
    satu entry per stasiun.
    """
    hasil = []
    for station_fields in json_data.get("data", []):
        meta, jam_data = parse_station(station_fields)
        hasil.append((meta, jam_data))
    return hasil


class FetchWorker(QThread):
    progress = Signal(str)
    selesai = Signal(bool, str, list)

    def __init__(self, resource, balai, provinsi, tgl_awal, tgl_akhir, parent=None):
        super().__init__(parent)
        self.resource = resource
        self.balai = balai
        self.provinsi = provinsi
        self.tgl_awal = tgl_awal
        self.tgl_akhir = tgl_akhir

    def run(self):
        try:
            config = load_config()
            session = build_session(config)

            hasil_semua = []
            error_log = []  # kumpulkan semua error per-tanggal, jangan hilang tertimpa
            mulai = datetime.strptime(self.tgl_awal, "%Y-%m-%d")
            akhir = datetime.strptime(self.tgl_akhir, "%Y-%m-%d")

            current = mulai
            while current <= akhir:
                tanggal_str = current.strftime("%Y-%m-%d")
                self.progress.emit(f"Mengambil data tanggal {tanggal_str}...")
                try:
                    json_data = fetch_daily(
                        session, tanggal_str, self.balai, self.provinsi, self.resource
                    )
                    hasil_parse = parse_response(json_data)
                    if not hasil_parse:
                        error_log.append(
                            f"{tanggal_str}: API sukses tapi 'data' kosong "
                            f"(status={json_data.get('status')}, code={json_data.get('code')})"
                        )
                    for meta, jam_data in hasil_parse:
                        hasil_semua.append((tanggal_str, self.resource, meta, jam_data))
                except Exception as e:
                    error_log.append(f"{tanggal_str}: {e}")
                current += timedelta(days=1)

            if hasil_semua:
                pesan = f"Berhasil mengambil {len(hasil_semua)} baris data."
                if error_log:
                    pesan += f" ({len(error_log)} tanggal bermasalah, lihat detail)"
            else:
                # 0 baris -> tampilkan penyebabnya, jangan cuma "0 baris data"
                contoh_error = "\n".join(error_log[:5]) if error_log else "(tidak ada detail error)"
                pesan = (
                    "Tidak ada data yang berhasil diambil (0 baris).\n\n"
                    f"Detail:\n{contoh_error}"
                )

            self.selesai.emit(True, pesan, hasil_semua)
        except Exception as e:
            self.selesai.emit(False, f"Gagal mengambil data:\n{e}", [])


class MonitoringPage(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.hasil_semua = []  # list of (tanggal, resource, meta, jam_data) hasil fetch terakhir
        self.worker = None

        layout = QVBoxLayout(self)
        layout.setContentsMargins(30, 20, 30, 20)
        layout.setSpacing(16)

        # ---- Judul ----
        title = QLabel("Monitoring Data")
        title.setFont(QFont("Arial", 16, QFont.Bold))
        title.setStyleSheet("color: #000000;")
        layout.addWidget(title)

        # ---- Form Filter ----
        form_frame = QWidget()
        form_frame.setStyleSheet("""
            QLabel { color: #000000; background-color: transparent; font-size: 12px; font-weight: bold; }
            QComboBox, QLineEdit, QDateEdit {
                background-color: #FFFFFF;
                color: #000000;
                border: 1px solid #A0A0A0;
                border-radius: 6px;
                padding: 6px 10px;
                font-size: 12px;
                min-height: 28px;
            }
            QComboBox QAbstractItemView {
                background-color: #FFFFFF;
                color: #000000;
                selection-background-color: #0077D4;
                selection-color: #FFFFFF;
            }
            QCalendarWidget QAbstractItemView {
                background-color: #FFFFFF;
                color: #000000;
                selection-background-color: #0077D4;
                selection-color: #FFFFFF;
            }
            QCalendarWidget QToolButton {
                color: #000000;
                background-color: transparent;
            }
            QCalendarWidget QMenu {
                background-color: #FFFFFF;
                color: #000000;
            }
            QCalendarWidget QSpinBox {
                background-color: #FFFFFF;
                color: #000000;
            }
        """)
        form_layout = QHBoxLayout(form_frame)
        form_layout.setSpacing(16)
        form_layout.setContentsMargins(0, 0, 0, 0)
        from PySide6.QtWidgets import QSizePolicy

        # Jenis Monitoring
        col1 = QVBoxLayout()
        col1.addWidget(QLabel("Jenis Monitoring"))
        self.combo_jenis = QComboBox()
        self.combo_jenis.addItems(list(RESOURCE_MAP.keys()))
        self.combo_jenis.setMinimumWidth(160)
        col1.addWidget(self.combo_jenis)
        form_layout.addLayout(col1, 2)

        # Balai (otomatis)
        col2 = QVBoxLayout()
        col2.addWidget(QLabel("Balai"))
        self.input_balai = QLineEdit("Balai III")
        self.input_balai.setMinimumWidth(90)
        col2.addWidget(self.input_balai)
        form_layout.addLayout(col2, 1)

        # Provinsi (otomatis)
        col3 = QVBoxLayout()
        col3.addWidget(QLabel("Provinsi"))
        self.input_provinsi = QLineEdit("Jawa Timur")
        self.input_provinsi.setMinimumWidth(90)
        col3.addWidget(self.input_provinsi)
        form_layout.addLayout(col3, 1)

        # Tanggal Mulai
        col4 = QVBoxLayout()
        col4.addWidget(QLabel("Tanggal Mulai"))
        self.date_awal = QDateEdit()
        self.date_awal.setCalendarPopup(True)
        self.date_awal.setDisplayFormat("yyyy-MM-dd")
        self.date_awal.setDate(QDate.currentDate().addDays(-1))
        self.date_awal.setMinimumWidth(130)
        col4.addWidget(self.date_awal)
        form_layout.addLayout(col4, 2)

        # Tanggal Akhir
        col5 = QVBoxLayout()
        col5.addWidget(QLabel("Tanggal Akhir"))
        self.date_akhir = QDateEdit()
        self.date_akhir.setCalendarPopup(True)
        self.date_akhir.setDisplayFormat("yyyy-MM-dd")
        self.date_akhir.setDate(QDate.currentDate())
        self.date_akhir.setMinimumWidth(130)
        col5.addWidget(self.date_akhir)
        form_layout.addLayout(col5, 2)

        calendar_style = """
            QCalendarWidget QAbstractItemView {
                background-color: #FFFFFF;
                color: #000000;
                selection-background-color: #0077D4;
                selection-color: #FFFFFF;
            }
            QCalendarWidget QToolButton { color: #000000; background-color: transparent; }
            QCalendarWidget QMenu { background-color: #FFFFFF; color: #000000; }
            QCalendarWidget QSpinBox { background-color: #FFFFFF; color: #000000; }
        """
        for date_edit in (self.date_awal, self.date_akhir):
            calendar = date_edit.calendarWidget()
            if calendar is not None:
                calendar.setStyleSheet(calendar_style)


        # Tombol Ambil Data
        col6 = QVBoxLayout()
        col6.addWidget(QLabel(""))  # spacer biar sejajar dengan input lain
        self.btn_ambil = QPushButton("Ambil Data")
        self.btn_ambil.setStyleSheet("""
            QPushButton {
                background-color: #0077D4; color: white; font-weight: bold;
                border: none; border-radius: 4px; padding: 6px 16px; font-size: 12px; min-height: 30px;
            }
            QPushButton:hover { background-color: #005FA3; }
            QPushButton:disabled { background-color: #A0C4E4; }
        """)
        self.btn_ambil.clicked.connect(self.ambil_data)
        col6.addWidget(self.btn_ambil)
        form_layout.addLayout(col6)
        # form_layout.addStretch()
        layout.addWidget(form_frame)

        # ---- Baris status + tombol export ----
        status_row = QHBoxLayout()
        self.label_status = QLabel("Belum ada data. Silakan pilih filter lalu klik 'Ambil Data'.")
        self.label_status.setStyleSheet("color: #555555; font-size: 12px;")
        status_row.addWidget(self.label_status)
        status_row.addStretch()

        self.btn_export = QPushButton("Export Excel")
        self.btn_export.setEnabled(False)
        self.btn_export.setStyleSheet("""
            QPushButton {
                background-color: #2ECC71; color: white; font-weight: bold;
                border: none; border-radius: 6px; padding: 8px 20px; font-size: 13px;
            }
            QPushButton:hover { background-color: #27AE60; }
            QPushButton:disabled { background-color: #A9DFBF; }
        """)
        self.btn_export.clicked.connect(self.export_excel)
        status_row.addWidget(self.btn_export)
        layout.addLayout(status_row)

        # ---- Tabel Hasil ----
        self.table = QTableWidget()
        self.table.setSelectionMode(QAbstractItemView.NoSelection)
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table.setFocusPolicy(Qt.NoFocus)
        self.table.setAlternatingRowColors(True)
        self.table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                gridline-color: #E0E0E0;
                border: 1px solid #D0D0D0;
                border-radius: 4px;
                selection-background-color: transparent;
                selection-color: #000000;
            }
            QTableWidget::item:alternate { background-color: #F9F9F9; }
            QTableWidget::item {
                color: #000000; font-size: 11px; padding-left: 6px;
            }
            QHeaderView::section {
                background-color: #F0F4F8; color: #000000; font-weight: bold;
                font-size: 11px; padding: 6px; border: none;
                border-bottom: 2px solid #D0D0D0;
            }
        """)
        self.table.verticalHeader().setVisible(False)
        layout.addWidget(self.table)

    def tampilkan_pesan(self, judul, pesan, jenis="info"):
        msg = QMessageBox(self)
        msg.setWindowTitle(judul)
        msg.setText(pesan)
        if jenis == "success":
            msg.setIcon(QMessageBox.Information)
        elif jenis == "warning":
            msg.setIcon(QMessageBox.Warning)
        elif jenis == "error":
            msg.setIcon(QMessageBox.Critical)
        else:
            msg.setIcon(QMessageBox.Information)
        msg.setStyleSheet("""
            QMessageBox { background-color: #FFFFFF; }
            QLabel { color: #000000; background-color: transparent; }
            QPushButton {
                color: #000000; background-color: #F0F4F8;
                border: 1px solid #A0A0A0; border-radius: 4px; padding: 5px 14px;
            }
            QPushButton:hover { background-color: #E0E6EA; }
        """)
        msg.exec()

    def ambil_data(self):
        jenis = self.combo_jenis.currentText()
        resource = RESOURCE_MAP[jenis]
        balai = self.input_balai.text().strip()
        provinsi = self.input_provinsi.text().strip()
        tgl_awal = self.date_awal.date().toString("yyyy-MM-dd")
        tgl_akhir = self.date_akhir.date().toString("yyyy-MM-dd")

        if not balai or not provinsi:
            self.tampilkan_pesan("Peringatan", "Balai dan Provinsi wajib diisi!", jenis="warning")
            return

        if self.date_awal.date() > self.date_akhir.date():
            self.tampilkan_pesan("Peringatan", "Tanggal mulai tidak boleh lebih besar dari tanggal akhir!", jenis="warning")
            return

        self.btn_ambil.setEnabled(False)
        self.btn_export.setEnabled(False)
        self.label_status.setText("Mengambil data, mohon tunggu...")

        self.worker = FetchWorker(resource, balai, provinsi, tgl_awal, tgl_akhir, self)
        self.worker.progress.connect(self.label_status.setText)
        self.worker.selesai.connect(self.on_fetch_selesai)
        self.worker.start()

    def on_fetch_selesai(self, sukses, pesan, hasil_semua):
        self.btn_ambil.setEnabled(True)

        self.label_status.setText(pesan.split("\n")[0])

        if not sukses:
            self.tampilkan_pesan("Gagal Mengambil Data", pesan, jenis="error")
            return

        if not hasil_semua:
            self.tampilkan_pesan("Data Kosong", pesan, jenis="warning")

        self.hasil_semua = hasil_semua
        self.render_table(hasil_semua)
        self.btn_export.setEnabled(bool(hasil_semua))

    # RENDER TABEL
    def render_table(self, hasil_semua):
        self.table.clear()
        self.table.setRowCount(0)

        if not hasil_semua:
            self.table.setColumnCount(0)
            return

        semua_jam = set()
        for _, _, _, jam_data in hasil_semua:
            semua_jam.update(jam_data.keys())
        semua_jam = sorted(semua_jam)

        headers = ["Tanggal", "Resource", "Station ID", "Station Name",
                   "Region", "Propinsi", "Kabupaten"] + semua_jam
        self.table.setColumnCount(len(headers))
        self.table.setHorizontalHeaderLabels(headers)

        for row_idx, (tanggal, resource, meta, jam_data) in enumerate(hasil_semua):
            self.table.insertRow(row_idx)
            fields = [
                tanggal, resource,
                meta.get("station_id", ""), meta.get("station_name", ""),
                meta.get("region", ""), meta.get("propinsi", ""), meta.get("kabupaten", ""),
            ]
            for col_idx, val in enumerate(fields):
                self.table.setItem(row_idx, col_idx, QTableWidgetItem(str(val)))

            for j_idx, jam in enumerate(semua_jam):
                entry = jam_data.get(jam)
                text = "" if entry is None else str(entry.get("value", ""))
                item = QTableWidgetItem(text)
                if entry and entry.get("status") == "late":
                    item.setForeground(Qt.red)
                self.table.setItem(row_idx, 7 + j_idx, item)

        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        for i in range(7):
            self.table.resizeColumnToContents(i)

    # EXPORT KE EXCEL
    def export_excel(self):
        if not self.hasil_semua:
            self.tampilkan_pesan("Peringatan", "Belum ada data untuk diexport.", jenis="warning")
            return

        tgl_awal = self.date_awal.date().toString("yyyy-MM-dd")
        tgl_akhir = self.date_akhir.date().toString("yyyy-MM-dd")
        default_name = f"Monitoring_{tgl_awal}_sd_{tgl_akhir}.xlsx"
        default_path = os.path.join(CURRENT_DIR, "hasil_export", default_name)

        filepath, _ = QFileDialog.getSaveFileName(
            self, "Simpan Hasil Export", default_path, "Excel Files (*.xlsx)"
        )
        if not filepath:
            return

        try:
            os.makedirs(os.path.dirname(filepath), exist_ok=True)
            self._tulis_excel(self.hasil_semua, filepath)
        except Exception as e:
            self.tampilkan_pesan("Gagal Export", f"Gagal menyimpan file Excel:\n{e}", jenis="error")
            return

        self.tampilkan_pesan("Berhasil", f"Data berhasil diexport ke:\n{filepath}", jenis="success")

    @staticmethod
    def _tulis_excel(hasil_semua, filepath):
        """Menulis hasil_semua (list of (tanggal, resource, meta, jam_data)) ke
        file Excel dengan 2 sheet, strukturnya sama seperti export_ke_excel()
        di test_monitor.py: sheet 'Monitoring' (ringkasan) + sheet
        'Riwayat Pengiriman' (detail per event kirim data)."""
        from openpyxl import Workbook

        semua_jam = set()
        for _, _, _, jam_data in hasil_semua:
            semua_jam.update(jam_data.keys())
        semua_jam = sorted(semua_jam)

        wb = Workbook()
        ws = wb.active
        ws.title = "Monitoring"
        header_row = ["Tanggal", "Resource", "Station ID", "Station Name",
                      "Region", "Propinsi", "Kabupaten"] + semua_jam
        ws.append(header_row)

        for tanggal, resource, meta, jam_data in hasil_semua:
            row = [
                tanggal, resource,
                meta.get("station_id"), meta.get("station_name"),
                meta.get("region"), meta.get("propinsi"), meta.get("kabupaten"),
            ]
            for j in semua_jam:
                entry = jam_data.get(j)
                row.append("" if entry is None else entry.get("value"))
            ws.append(row)

        for i, h in enumerate(header_row, start=1):
            col_letter = ws.cell(row=1, column=i).column_letter
            ws.column_dimensions[col_letter].width = max(10, len(str(h)) + 2)
        ws.freeze_panes = "A2"

        ws2 = wb.create_sheet("Riwayat Pengiriman")
        header_riwayat = [
            "Tanggal", "Station ID", "Station Name", "Jam", "Status",
            "Delta (menit)", "Timestamp Data (seharusnya)", "Timestamp Dikirim (aktual)",
        ]
        ws2.append(header_riwayat)

        for tanggal, resource, meta, jam_data in hasil_semua:
            for jam, entry in sorted(jam_data.items()):
                raw_list = entry.get("raw") or []
                if not raw_list:
                    ws2.append([tanggal, meta.get("station_id"), meta.get("station_name"),
                                jam, entry.get("status"), "", "", ""])
                else:
                    for raw_entry in raw_list:
                        ws2.append([
                            tanggal, meta.get("station_id"), meta.get("station_name"),
                            jam, entry.get("status"),
                            raw_entry.get("delta_min"),
                            raw_entry.get("timestamp_data"),
                            raw_entry.get("timestamp_sent_data"),
                        ])

        for i, h in enumerate(header_riwayat, start=1):
            col_letter = ws2.cell(row=1, column=i).column_letter
            ws2.column_dimensions[col_letter].width = max(12, len(str(h)) + 2)
        ws2.freeze_panes = "A2"

        wb.save(filepath)